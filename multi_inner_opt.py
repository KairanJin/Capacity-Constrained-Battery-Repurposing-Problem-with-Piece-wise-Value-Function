from __future__ import annotations

import math
from pathlib import Path
from typing import Dict, List

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from config import Config
from data_generator import generate_cells
from heuristics.rrp_ga import solve_rrp_ga
from heuristics.rrp_kmeans import solve_rrp_kmeans
from heuristics.rrp_kmeans_vns import solve_rrp_kmeans_vns
from heuristics.rrp_grasp import solve_rrp_grasp
from heuristics.rrp_column_generation import solve_rrp_column_generation
from heuristics.rrp_ms_kmeans_vns import solve_rrp_ms_kmeans_vns
from utils import safe_div, summarize_solution


METHODS = [
    "RRP_KMEANS",
    "RRP_KMEANS_VNS",
    "RRP_MS_KMEANS_VNS",
    # "RRP_GRASP",
    "RRP_GA",
    "RRP_COLUMN_GENERATION",
]

METHOD_LABELS = {
    "RRP_KMEANS": "KMeans",
    "RRP_KMEANS_VNS": "KMeans_VNS",
    "RRP_MS_KMEANS_VNS": "MS_KMeans_VNS",
    "RRP_GRASP": "GRASP",
    "RRP_GA": "GA",
    "RRP_COLUMN_GENERATION": "ColumnGeneration",
}


def format_tier_distribution(tier_counts: dict) -> str:
    return (
        f"P1:{tier_counts.get('P1', 0)}, "
        f"P2:{tier_counts.get('P2', 0)}, "
        f"P3:{tier_counts.get('P3', 0)}, "
        f"P0:{tier_counts.get('P0', 0)}"
    )


def enrich_result(result: dict, X: np.ndarray, cfg: Config) -> dict:
    summary = summarize_solution(
        X=X,
        groups=result["groups"],
        K=cfg.problem.K,
        w=cfg.problem.w,
        lambda_penalty=cfg.problem.lambda_penalty,
        theta1=cfg.problem.theta1,
        theta2=cfg.problem.theta2,
        theta3=cfg.problem.theta3,
        P1=cfg.problem.P1,
        P2=cfg.problem.P2,
        P3=cfg.problem.P3,
    )

    total_cells = X.shape[0]
    used_cells = cfg.problem.K * summary["n_packs"]
    utilization_rate = safe_div(used_cells, total_cells)

    result["reward"] = summary["total_reward"]
    result["n_packs"] = summary["n_packs"]
    result["avg_delta"] = summary["avg_delta"]
    result["avg_phi"] = summary["avg_phi"]
    result["reward_per_pack"] = summary["reward_per_pack"]
    result["positive_pack_ratio"] = summary["positive_pack_ratio"]
    result["tier_counts"] = summary["tier_counts"]
    result["tier_distribution"] = format_tier_distribution(summary["tier_counts"])
    result["utilization_rate"] = utilization_rate
    return result


def make_empty_result(method: str, X: np.ndarray) -> dict:
    return {
        "method": method,
        "groups": [],
        "leftover": list(range(X.shape[0])),
        "reward": 0.0,
        "n_packs": 0,
        "avg_delta": 0.0,
        "avg_phi": 0.0,
        "runtime": 0.0,
        "reward_per_pack": 0.0,
        "positive_pack_ratio": 0.0,
        "tier_counts": {"P1": 0, "P2": 0, "P3": 0, "P0": 0},
        "tier_distribution": "P1:0, P2:0, P3:0, P0:0",
        "n_columns": float("nan"),
    }


def standardize_cells(X_raw: np.ndarray, cfg: Config) -> np.ndarray:
    """
    将原始电芯数据标准化为算法使用的 X。

    假设 generate_cells 输出的两列分别为：
    第 0 列：capacity C
    第 1 列：resistance R

    标准化方式：
    C_tilde = (C - mu_C) / sigma_C
    R_tilde = (R - mu_R) / sigma_R
    """
    X_raw = np.asarray(X_raw, dtype=float)

    X_std = np.column_stack(
        [
            (X_raw[:, 0] - cfg.data.mu_C) / cfg.data.sigma_C,
            (X_raw[:, 1] - cfg.data.mu_R) / cfg.data.sigma_R,
        ]
    )

    return X_std


def solve_one_method(method: str, X: np.ndarray, cfg: Config, seed: int) -> dict:
    k_t = min(cfg.problem.k_max, X.shape[0] // cfg.problem.K)
    if X.shape[0] < cfg.problem.K or k_t <= 0:
        return enrich_result(make_empty_result(method, X), X, cfg)

    if method == "RRP_KMEANS":
        res = solve_rrp_kmeans(
            X=X,
            K=cfg.problem.K,
            k_t=k_t,
            delta_bar=cfg.problem.delta_bar,
            L1=cfg.kmeans.L1,
            L2=cfg.kmeans.L2,
            tol=cfg.kmeans.tol,
            w=cfg.problem.w,
            lambda_penalty=cfg.problem.lambda_penalty,
            theta1=cfg.problem.theta1,
            theta2=cfg.problem.theta2,
            theta3=cfg.problem.theta3,
            P1=cfg.problem.P1,
            P2=cfg.problem.P2,
            P3=cfg.problem.P3,
            seed=seed,
        )

    elif method == "RRP_KMEANS_VNS":
        res = solve_rrp_kmeans_vns(
            X=X,
            K=cfg.problem.K,
            k_t=k_t,
            delta_bar=cfg.problem.delta_bar,
            L1=cfg.vns.L1,
            tol=cfg.vns.tol,
            max_vns_iter=cfg.vns.max_vns_iter,
            max_no_improve=cfg.vns.max_no_improve,
            w=cfg.problem.w,
            lambda_penalty=cfg.problem.lambda_penalty,
            theta1=cfg.problem.theta1,
            theta2=cfg.problem.theta2,
            theta3=cfg.problem.theta3,
            P1=cfg.problem.P1,
            P2=cfg.problem.P2,
            P3=cfg.problem.P3,
            seed=seed,
            pack_candidate_limit=cfg.vns.pack_candidate_limit,
            partner_limit=cfg.vns.partner_limit,
            cell_candidate_limit=cfg.vns.cell_candidate_limit,
            leftover_candidate_limit=cfg.vns.leftover_candidate_limit,
            destroy_size=cfg.vns.destroy_size,
        )

    elif method == "RRP_MS_KMEANS_VNS":
        res = solve_rrp_ms_kmeans_vns(
            X=X,
            K=cfg.problem.K,
            k_t=k_t,
            delta_bar=cfg.problem.delta_bar,
            w=cfg.problem.w,
            lambda_penalty=cfg.problem.lambda_penalty,
            theta1=cfg.problem.theta1,
            theta2=cfg.problem.theta2,
            theta3=cfg.problem.theta3,
            P1=cfg.problem.P1,
            P2=cfg.problem.P2,
            P3=cfg.problem.P3,
            seed=seed,
        )

    elif method == "RRP_GRASP":
        # solve_rrp_grasp.py 中已经设置了默认参数：
        # n_starts=20, rcl_size=4, max_group_attempts=200, max_local_iter=30 等。
        # 这里先只传公共问题参数，避免依赖 config.py 中不存在的 cfg.grasp。
        res = solve_rrp_grasp(
            X=X,
            K=cfg.problem.K,
            k_t=k_t,
            delta_bar=cfg.problem.delta_bar,
            w=cfg.problem.w,
            lambda_penalty=cfg.problem.lambda_penalty,
            theta1=cfg.problem.theta1,
            theta2=cfg.problem.theta2,
            theta3=cfg.problem.theta3,
            P1=cfg.problem.P1,
            P2=cfg.problem.P2,
            P3=cfg.problem.P3,
            seed=seed,
        )

    elif method == "RRP_GA":
        res = solve_rrp_ga(
            X=X,
            K=cfg.problem.K,
            k_t=k_t,
            delta_bar=cfg.problem.delta_bar,
            w=cfg.problem.w,
            lambda_penalty=cfg.problem.lambda_penalty,
            theta1=cfg.problem.theta1,
            theta2=cfg.problem.theta2,
            theta3=cfg.problem.theta3,
            P1=cfg.problem.P1,
            P2=cfg.problem.P2,
            P3=cfg.problem.P3,
            seed=seed,
            population_size=cfg.ga.population_size,
            n_generations=cfg.ga.n_generations,
            tournament_size=cfg.ga.tournament_size,
            crossover_prob=cfg.ga.crossover_prob,
            mutation_prob=cfg.ga.mutation_prob,
            destroy_size=cfg.ga.destroy_size,
            local_search_prob=cfg.ga.local_search_prob,
            elitism_size=cfg.ga.elitism_size,
            group_candidate_limit=cfg.ga.group_candidate_limit,
            cell_candidate_limit=cfg.ga.cell_candidate_limit,
            leftover_candidate_limit=cfg.ga.leftover_candidate_limit,
        )

    elif method == "RRP_COLUMN_GENERATION":
        # solve_rrp_column_generation.py 中已经设置了默认参数：
        # max_cg_iter=30, init_n_starts=30, pricing_n_seeds=40 等。
        # 这里先只传公共问题参数，避免依赖 config.py 中不存在的 cfg.column_generation。
        res = solve_rrp_column_generation(
            X=X,
            K=cfg.problem.K,
            k_t=k_t,
            delta_bar=cfg.problem.delta_bar,
            w=cfg.problem.w,
            lambda_penalty=cfg.problem.lambda_penalty,
            theta1=cfg.problem.theta1,
            theta2=cfg.problem.theta2,
            theta3=cfg.problem.theta3,
            P1=cfg.problem.P1,
            P2=cfg.problem.P2,
            P3=cfg.problem.P3,
            seed=seed,
            # 如果本地装了 Gurobi，默认会尝试 Gurobi pricing；
            # 没有 Gurobi 时，rrp_column_generation.py 内部会自动 fallback。
            # 若想强制只用 PuLP/CBC + local enumeration，可改为 False。
            use_gurobi_pricing=True,
        )

    else:
        raise ValueError(f"Unknown method: {method}")

    return enrich_result(res, X, cfg)


def run_multi_round_experiment(
    n_rounds: int = 20,
    arrivals_per_round: int = 130,
    output_excel_path: str = "rrp_multi_round_output.xlsx",
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    多轮重组逻辑：
    - 每轮新到 arrivals_per_round 个电芯
    - 不考虑报废
    - 每种算法维护自己的 leftover 池
    - 每轮的 pool = 上一轮 leftover + 本轮新到电芯
    - 每轮重组后的 leftover 进入下一轮

    重要修改：
    - generate_cells 生成的是原始尺度下的电芯数据
    - 在进入算法前，需要统一标准化
    - pool_by_method 中保存的 leftover 也是标准化后的数据
    """
    cfg = Config()
    base_seed = cfg.experiment.base_seed

    pool_by_method: Dict[str, np.ndarray | None] = {m: None for m in METHODS}
    cumulative_reward_by_method: Dict[str, float] = {m: 0.0 for m in METHODS}

    detail_rows: List[dict] = []

    for round_idx in range(1, n_rounds + 1):
        round_seed = base_seed + round_idx

        # =====================================================
        # 1) 生成原始电芯数据
        # =====================================================
        new_cells_raw = generate_cells(
            n_cells=arrivals_per_round,
            mu_C=cfg.data.mu_C,
            sigma_C=cfg.data.sigma_C,
            mu_R=cfg.data.mu_R,
            sigma_R=cfg.data.sigma_R,
            seed=round_seed,
        )

        # =====================================================
        # 2) 标准化后再进入算法
        # =====================================================
        new_cells = standardize_cells(new_cells_raw, cfg)

        # 只在第一轮打印一次标准化检查
        if round_idx == 1:
            print("\n=== Standardization Check ===")
            print("raw new_cells mean =", new_cells_raw.mean(axis=0))
            print("raw new_cells std  =", new_cells_raw.std(axis=0))
            print("std new_cells mean =", new_cells.mean(axis=0))
            print("std new_cells std  =", new_cells.std(axis=0))
            print("std new_cells min  =", new_cells.min(axis=0))
            print("std new_cells max  =", new_cells.max(axis=0))
            print("delta_bar =", cfg.problem.delta_bar)
            print("theta1, theta2, theta3 =", cfg.problem.theta1, cfg.problem.theta2, cfg.problem.theta3)

        for method_idx, method in enumerate(METHODS):
            # =====================================================
            # 3) 构造当前算法本轮的 pool
            # 注意：pool_by_method 中存储的也是标准化后的 leftover
            # =====================================================
            if pool_by_method[method] is None or pool_by_method[method].shape[0] == 0:
                pool_before = new_cells.copy()
            else:
                pool_before = np.vstack([pool_by_method[method], new_cells])

            print(
                f"[Round {round_idx}/{n_rounds}] Start {method} | pool size = {pool_before.shape[0]}"
            )

            # =====================================================
            # 4) 求解当前方法
            # =====================================================
            result = solve_one_method(
                method=method,
                X=pool_before,
                cfg=cfg,
                seed=round_seed * 100 + method_idx,
            )

            # =====================================================
            # 5) 保存 leftover 进入下一轮
            # =====================================================
            leftover_idx = result["leftover"]
            if len(leftover_idx) > 0:
                next_pool = pool_before[leftover_idx].copy()
            else:
                next_pool = np.empty((0, pool_before.shape[1]), dtype=pool_before.dtype)

            pool_by_method[method] = next_pool
            cumulative_reward_by_method[method] += result["reward"]

            tier_counts = result.get("tier_counts", {}) or {}
            used_cells = cfg.problem.K * result["n_packs"]

            detail_rows.append(
                {
                    "round": round_idx,
                    "method": method,
                    "method_label": METHOD_LABELS[method],
                    "new_cells": arrivals_per_round,
                    "pool_before": int(pool_before.shape[0]),
                    "reward": float(result["reward"]),
                    "cumulative_reward": float(cumulative_reward_by_method[method]),
                    "runtime": float(result["runtime"]),
                    "n_packs": int(result["n_packs"]),
                    "used_cells": int(used_cells),
                    "leftover_cells": int(len(result["leftover"])),
                    "utilization_rate": float(result["utilization_rate"]),
                    "avg_delta": float(result["avg_delta"]),
                    "avg_phi": float(result["avg_phi"]),
                    "reward_per_pack": float(result["reward_per_pack"]),
                    "positive_pack_ratio": float(result["positive_pack_ratio"]),
                    "P1_packs": int(tier_counts.get("P1", 0)),
                    "P2_packs": int(tier_counts.get("P2", 0)),
                    "P3_packs": int(tier_counts.get("P3", 0)),
                    "P0_packs": int(tier_counts.get("P0", 0)),
                    "P1_cells": int(tier_counts.get("P1", 0) * cfg.problem.K),
                    "P2_cells": int(tier_counts.get("P2", 0) * cfg.problem.K),
                    "P3_cells": int(tier_counts.get("P3", 0) * cfg.problem.K),
                    "P0_cells": int(tier_counts.get("P0", 0) * cfg.problem.K),
                    "tier_distribution": result["tier_distribution"],
                    "n_columns": result.get("n_columns", float("nan")),
                }
            )

    detail_df = pd.DataFrame(detail_rows)

    summary_df = (
        detail_df.groupby(["method", "method_label"], as_index=False)
        .agg(
            total_reward=("reward", "sum"),
            final_cumulative_reward=("cumulative_reward", "max"),
            total_runtime=("runtime", "sum"),
            avg_runtime=("runtime", "mean"),
            total_packs=("n_packs", "sum"),
            total_used_cells=("used_cells", "sum"),
            final_leftover=("leftover_cells", "last"),
            P1_packs_total=("P1_packs", "sum"),
            P2_packs_total=("P2_packs", "sum"),
            P3_packs_total=("P3_packs", "sum"),
            P0_packs_total=("P0_packs", "sum"),
            P1_cells_total=("P1_cells", "sum"),
            P2_cells_total=("P2_cells", "sum"),
            P3_cells_total=("P3_cells", "sum"),
            P0_cells_total=("P0_cells", "sum"),
        )
        .sort_values("method")
        .reset_index(drop=True)
    )

    export_results_to_excel(
        detail_df=detail_df,
        summary_df=summary_df,
        output_path=output_excel_path,
        n_rounds=n_rounds,
    )

    return detail_df, summary_df


def _apply_header_style(ws, row: int = 1):
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E1F2")
    for cell in ws[row]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=thin)


def _set_col_widths(ws, widths: dict):
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _write_dataframe(ws, df: pd.DataFrame, start_row: int = 1, start_col: int = 1):
    headers = list(df.columns)
    for j, h in enumerate(headers, start=start_col):
        ws.cell(start_row, j, h)
    for i, (_, row) in enumerate(df.iterrows(), start=start_row + 1):
        for j, h in enumerate(headers, start=start_col):
            ws.cell(i, j, row[h])


def create_excel_template_only(
    output_path: str = "rrp_multi_round_template.xlsx",
    n_rounds: int = 20,
):
    """
    只生成模板，不填实际实验数据。
    后续你可以把代码跑出来的明细粘贴到 Round_Results 工作表。
    """
    empty_detail = pd.DataFrame(
        columns=[
            "round", "method", "method_label", "new_cells", "pool_before", "reward",
            "cumulative_reward", "runtime", "n_packs", "used_cells", "leftover_cells",
            "utilization_rate", "avg_delta", "avg_phi", "reward_per_pack",
            "positive_pack_ratio", "P1_packs", "P2_packs", "P3_packs", "P0_packs",
            "P1_cells", "P2_cells", "P3_cells", "P0_cells", "tier_distribution", "n_columns",
        ]
    )
    empty_summary = pd.DataFrame(
        {
            "method": METHODS,
            "method_label": [METHOD_LABELS[m] for m in METHODS],
        }
    )
    export_results_to_excel(
        detail_df=empty_detail,
        summary_df=empty_summary,
        output_path=output_path,
        n_rounds=n_rounds,
    )


def export_results_to_excel(
    detail_df: pd.DataFrame,
    summary_df: pd.DataFrame,
    output_path: str,
    n_rounds: int = 20,
):
    wb = Workbook()

    # Remove default
    default_ws = wb.active
    wb.remove(default_ws)

    # =========================
    # README
    # =========================
    ws = wb.create_sheet("README")
    ws["A1"] = "20轮重组实验 Excel 模板"
    ws["A1"].font = Font(size=14, bold=True)
    ws["A3"] = "说明："
    ws["A4"] = "1. Round_Results 为逐轮明细表，后续可直接按相同列顺序粘贴数据。"
    ws["A5"] = "2. ChartData_CumReward：五种算法每轮累计收益绘图数据。"
    ws["A6"] = "3. ChartData_TierCells：五种算法20轮中各类电芯重组总数量绘图数据。"
    ws["A7"] = "4. ChartData_Runtime：五种算法每轮运行时间绘图数据。"
    ws["A8"] = "5. Charts 工作表内已预置三张图。"
    ws["A10"] = "方法列表："
    for i, method in enumerate(METHODS, start=11):
        ws.cell(i, 1, method)
        ws.cell(i, 2, METHOD_LABELS[method])
    ws.column_dimensions["A"].width = 60
    ws.column_dimensions["B"].width = 25

    # =========================
    # Round_Results
    # =========================
    ws = wb.create_sheet("Round_Results")
    if detail_df.empty:
        rows = []
        for r in range(1, n_rounds + 1):
            for method in METHODS:
                rows.append(
                    {
                        "round": r,
                        "method": method,
                        "method_label": METHOD_LABELS[method],
                        "new_cells": None,
                        "pool_before": None,
                        "reward": None,
                        "cumulative_reward": None,
                        "runtime": None,
                        "n_packs": None,
                        "used_cells": None,
                        "leftover_cells": None,
                        "utilization_rate": None,
                        "avg_delta": None,
                        "avg_phi": None,
                        "reward_per_pack": None,
                        "positive_pack_ratio": None,
                        "P1_packs": None,
                        "P2_packs": None,
                        "P3_packs": None,
                        "P0_packs": None,
                        "P1_cells": None,
                        "P2_cells": None,
                        "P3_cells": None,
                        "P0_cells": None,
                        "tier_distribution": None,
                        "n_columns": None,
                    }
                )
        detail_df_to_write = pd.DataFrame(rows)
    else:
        detail_df_to_write = detail_df.copy()

    _write_dataframe(ws, detail_df_to_write)
    _apply_header_style(ws, 1)
    ws.freeze_panes = "A2"
    _set_col_widths(
        ws,
        {
            1: 8, 2: 24, 3: 20, 4: 12, 5: 12, 6: 14, 7: 18, 8: 12, 9: 10, 10: 12,
            11: 14, 12: 14, 13: 12, 14: 12, 15: 16, 16: 18, 17: 10, 18: 10, 19: 10,
            20: 10, 21: 10, 22: 10, 23: 10, 24: 10, 25: 28, 26: 12,
        },
    )

    pct_cols = ["L", "P"]
    for col in pct_cols:
        for row in range(2, ws.max_row + 1):
            ws[f"{col}{row}"].number_format = "0.0%"

    decimal_cols = ["F", "G", "H", "M", "N", "O"]
    for col in decimal_cols:
        for row in range(2, ws.max_row + 1):
            ws[f"{col}{row}"].number_format = "0.000"

    # =========================
    # Method_Summary
    # =========================
    ws = wb.create_sheet("Method_Summary")
    summary_headers = [
        "method", "method_label", "total_reward", "final_cumulative_reward",
        "total_runtime", "avg_runtime", "total_packs", "total_used_cells",
        "final_leftover", "P1_cells_total", "P2_cells_total", "P3_cells_total", "P0_cells_total",
    ]
    for j, h in enumerate(summary_headers, start=1):
        ws.cell(1, j, h)
    _apply_header_style(ws, 1)

    for i, method in enumerate(METHODS, start=2):
        ws.cell(i, 1, method)
        ws.cell(i, 2, METHOD_LABELS[method])
        ws.cell(i, 3, f'=SUMIFS(Round_Results!$F:$F,Round_Results!$B:$B,$A{i})')
        ws.cell(i, 4, f'=MAXIFS(Round_Results!$G:$G,Round_Results!$B:$B,$A{i})')
        ws.cell(i, 5, f'=SUMIFS(Round_Results!$H:$H,Round_Results!$B:$B,$A{i})')
        ws.cell(i, 6, f'=AVERAGEIFS(Round_Results!$H:$H,Round_Results!$B:$B,$A{i})')
        ws.cell(i, 7, f'=SUMIFS(Round_Results!$I:$I,Round_Results!$B:$B,$A{i})')
        ws.cell(i, 8, f'=SUMIFS(Round_Results!$J:$J,Round_Results!$B:$B,$A{i})')
        ws.cell(i, 9, f'=MAXIFS(Round_Results!$K:$K,Round_Results!$B:$B,$A{i},Round_Results!$A:$A,{n_rounds})')
        ws.cell(i, 10, f'=SUMIFS(Round_Results!$U:$U,Round_Results!$B:$B,$A{i})')
        ws.cell(i, 11, f'=SUMIFS(Round_Results!$V:$V,Round_Results!$B:$B,$A{i})')
        ws.cell(i, 12, f'=SUMIFS(Round_Results!$W:$W,Round_Results!$B:$B,$A{i})')
        ws.cell(i, 13, f'=SUMIFS(Round_Results!$X:$X,Round_Results!$B:$B,$A{i})')

    _set_col_widths(
        ws,
        {
            1: 24, 2: 20, 3: 18, 4: 22, 5: 16, 6: 14, 7: 14,
            8: 16, 9: 14, 10: 14, 11: 14, 12: 14, 13: 14,
        },
    )

    # =========================
    # ChartData_CumReward
    # =========================
    ws = wb.create_sheet("ChartData_CumReward")
    ws["A1"] = "round"
    for j, method in enumerate(METHODS, start=2):
        ws.cell(1, j, method)
    for i in range(2, n_rounds + 2):
        ws.cell(i, 1, i - 1)
        for j, method in enumerate(METHODS, start=2):
            ws.cell(
                i,
                j,
                f'=SUMIFS(Round_Results!$F:$F,Round_Results!$B:$B,{get_column_letter(j)}$1,Round_Results!$A:$A,"<="&$A{i})',
            )
    _apply_header_style(ws, 1)
    _set_col_widths(ws, {1: 10, 2: 18, 3: 18, 4: 18, 5: 18, 6: 24})

    # =========================
    # ChartData_TierCells
    # =========================
    ws = wb.create_sheet("ChartData_TierCells")
    headers = ["method", "P1_cells_total", "P2_cells_total", "P3_cells_total", "P0_cells_total"]
    for j, h in enumerate(headers, start=1):
        ws.cell(1, j, h)
    for i, method in enumerate(METHODS, start=2):
        ws.cell(i, 1, method)
        ws.cell(i, 2, f'=SUMIFS(Round_Results!$U:$U,Round_Results!$B:$B,$A{i})')
        ws.cell(i, 3, f'=SUMIFS(Round_Results!$V:$V,Round_Results!$B:$B,$A{i})')
        ws.cell(i, 4, f'=SUMIFS(Round_Results!$W:$W,Round_Results!$B:$B,$A{i})')
        ws.cell(i, 5, f'=SUMIFS(Round_Results!$X:$X,Round_Results!$B:$B,$A{i})')
    _apply_header_style(ws, 1)
    _set_col_widths(ws, {1: 24, 2: 16, 3: 16, 4: 16, 5: 16})

    # =========================
    # ChartData_Runtime
    # =========================
    ws = wb.create_sheet("ChartData_Runtime")
    ws["A1"] = "round"
    for j, method in enumerate(METHODS, start=2):
        ws.cell(1, j, method)
    for i in range(2, n_rounds + 2):
        ws.cell(i, 1, i - 1)
        for j, method in enumerate(METHODS, start=2):
            ws.cell(
                i,
                j,
                f'=SUMIFS(Round_Results!$H:$H,Round_Results!$B:$B,{get_column_letter(j)}$1,Round_Results!$A:$A,$A{i})',
            )
    _apply_header_style(ws, 1)
    _set_col_widths(ws, {1: 10, 2: 18, 3: 18, 4: 18, 5: 18, 6: 24})

    # =========================
    # Charts
    # =========================
    ws = wb.create_sheet("Charts")

    # Chart 1: cumulative reward
    chart1 = LineChart()
    chart1.title = "五种算法每轮累计收益对比"
    chart1.y_axis.title = "累计收益"
    chart1.x_axis.title = "轮次"
    chart1.height = 9
    chart1.width = 18
    data = Reference(
        wb["ChartData_CumReward"],
        min_col=2,
        max_col=1 + len(METHODS),
        min_row=1,
        max_row=n_rounds + 1,
    )
    cats = Reference(
        wb["ChartData_CumReward"],
        min_col=1,
        min_row=2,
        max_row=n_rounds + 1,
    )
    chart1.add_data(data, titles_from_data=True)
    chart1.set_categories(cats)
    chart1.style = 2
    ws.add_chart(chart1, "A1")

    # Chart 2: total tier cells
    chart2 = BarChart()
    chart2.type = "bar"
    chart2.style = 10
    chart2.title = "五种算法在20轮中的各类电芯重组总数量对比"
    chart2.y_axis.title = "算法"
    chart2.x_axis.title = "总重组电芯数量"
    chart2.height = 10
    chart2.width = 18
    data = Reference(
        wb["ChartData_TierCells"],
        min_col=2,
        max_col=5,
        min_row=1,
        max_row=1 + len(METHODS),
    )
    cats = Reference(
        wb["ChartData_TierCells"],
        min_col=1,
        min_row=2,
        max_row=1 + len(METHODS),
    )
    chart2.add_data(data, titles_from_data=True)
    chart2.set_categories(cats)
    ws.add_chart(chart2, "A20")

    # Chart 3: runtime per round
    chart3 = LineChart()
    chart3.title = "五种算法每轮运行时间对比"
    chart3.y_axis.title = "运行时间"
    chart3.x_axis.title = "轮次"
    chart3.height = 9
    chart3.width = 18
    data = Reference(
        wb["ChartData_Runtime"],
        min_col=2,
        max_col=1 + len(METHODS),
        min_row=1,
        max_row=n_rounds + 1,
    )
    cats = Reference(
        wb["ChartData_Runtime"],
        min_col=1,
        min_row=2,
        max_row=n_rounds + 1,
    )
    chart3.add_data(data, titles_from_data=True)
    chart3.set_categories(cats)
    chart3.style = 2
    ws.add_chart(chart3, "T1")

    ws.sheet_view.showGridLines = False

    # number formats
    for ws_name in ["Method_Summary", "ChartData_CumReward", "ChartData_Runtime", "ChartData_TierCells"]:
        ws = wb[ws_name]
        for row in ws.iter_rows(min_row=2):
            for cell in row[1:]:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    cell.number_format = "0.000"

    output_path = str(Path(output_path))
    wb.save(output_path)


def main():
    # 1) 生成可直接复用的空模板
    create_excel_template_only("rrp_multi_round_template.xlsx", n_rounds=20)

    # 2) 实际跑实验，并输出结果到 Excel
    detail_df, summary_df = run_multi_round_experiment(
        n_rounds=10,
        arrivals_per_round=400,
        output_excel_path="rrp_multi_round_output.xlsx",
    )

    print("\n=== Multi-round Detail (head) ===")
    with pd.option_context("display.max_columns", None, "display.width", 260):
        print(detail_df.head(15).to_string(index=False))

    print("\n=== Multi-round Summary ===")
    with pd.option_context("display.max_columns", None, "display.width", 260):
        print(summary_df.to_string(index=False))


if __name__ == "__main__":
    main()