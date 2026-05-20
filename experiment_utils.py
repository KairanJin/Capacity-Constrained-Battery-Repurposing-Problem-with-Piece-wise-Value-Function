"""
数值实验工具模块：统计检验 + Excel 结果生成
"""
import os
import numpy as np
import pandas as pd
from scipy import stats
from itertools import combinations
from datetime import datetime

# =========================================================
# 统计检验
# =========================================================

def paired_ttest(results_df, method_a, method_b, metric="reward"):
    """
    配对 t 检验（同实例同种子配对）。
    返回 (t_statistic, p_value, mean_a, mean_b, mean_diff)
    """
    mask_a = results_df["method"] == method_a
    mask_b = results_df["method"] == method_b
    # 按 (instance_id, seed) 配对
    df_a = results_df[mask_a][["instance_id", "seed", metric]].copy()
    df_b = results_df[mask_b][["instance_id", "seed", metric]].copy()
    merged = df_a.merge(df_b, on=["instance_id", "seed"], how="inner", suffixes=("_a", "_b"))
    if len(merged) < 2:
        return None
    vals_a = merged[metric + "_a"].values.astype(float)
    vals_b = merged[metric + "_b"].values.astype(float)
    t_stat, p_val = stats.ttest_rel(vals_a, vals_b)
    return {
        "method_a": method_a,
        "method_b": method_b,
        "metric": metric,
        "mean_a": float(np.mean(vals_a)),
        "mean_b": float(np.mean(vals_b)),
        "mean_diff": float(np.mean(vals_a - vals_b)),
        "t_statistic": float(t_stat),
        "p_value": float(p_val),
        "n_pairs": len(merged),
        "significant_005": p_val < 0.05,
        "significant_001": p_val < 0.01,
    }


def cohens_d(results_df, method_a, method_b, metric="reward"):
    """
    Cohen's d 效应量。
    d = (mean_a - mean_b) / pooled_std
    """
    mask_a = results_df["method"] == method_a
    mask_b = results_df["method"] == method_b
    df_a = results_df[mask_a][["instance_id", "seed", metric]].copy()
    df_b = results_df[mask_b][["instance_id", "seed", metric]].copy()
    merged = df_a.merge(df_b, on=["instance_id", "seed"], how="inner", suffixes=("_a", "_b"))
    if len(merged) < 2:
        return None
    vals_a = merged[metric + "_a"].values.astype(float)
    vals_b = merged[metric + "_b"].values.astype(float)
    mean_diff = np.mean(vals_a) - np.mean(vals_b)
    pooled_std = np.sqrt((np.std(vals_a, ddof=1) ** 2 + np.std(vals_b, ddof=1) ** 2) / 2)
    if pooled_std < 1e-12:
        return 0.0
    return float(mean_diff / pooled_std)


def anova_test(results_df, methods, metric="reward"):
    """
    单因素 ANOVA（F 检验），检验多算法整体差异。
    """
    groups = []
    for m in methods:
        vals = results_df[results_df["method"] == m][metric].dropna().values.astype(float)
        if len(vals) >= 2:
            groups.append(vals)
    if len(groups) < 2:
        return None
    f_stat, p_val = stats.f_oneway(*groups)
    return {
        "metric": metric,
        "f_statistic": float(f_stat),
        "p_value": float(p_val),
        "n_groups": len(groups),
        "total_samples": sum(len(g) for g in groups),
        "significant_005": p_val < 0.05,
    }


def tukey_hsd_test(results_df, methods, metric="reward"):
    """
    Tukey HSD 事后检验（成对比较）。
    scipy >= 1.11 API: tukey_hsd(*samples)
    """
    groups_data = {}
    for m in methods:
        vals = results_df[results_df["method"] == m][metric].dropna().values.astype(float)
        if len(vals) >= 2:
            groups_data[m] = vals
    if len(groups_data) < 2:
        return []

    group_names = list(groups_data.keys())
    samples = [groups_data[m] for m in group_names]
    tukey_result = stats.tukey_hsd(*samples)

    results = []
    ci = tukey_result.confidence_interval()
    for i in range(len(group_names)):
        for j in range(i + 1, len(group_names)):
            mean_diff = float(np.mean(samples[i]) - np.mean(samples[j]))
            results.append({
                "method_a": group_names[i],
                "method_b": group_names[j],
                "mean_diff": mean_diff,
                "p_adj": float(tukey_result.pvalue[i, j]),
                "reject": bool(tukey_result.pvalue[i, j] < 0.05),
                "conf_low": float(ci.low[i, j]),
                "conf_high": float(ci.high[i, j]),
            })
    return results


def run_all_statistical_tests(results_df, methods, ref_method="RRP_KMEANS_VNS", metric="reward"):
    """
    运行全部统计检验，返回 dict:
      - paired_ttests: list of dict (ref_method vs each other)
      - cohens_ds: dict {other_method: d_value}
      - anova: dict
      - tukey: list of dict
    """
    ttests = []
    ds = {}
    for m in methods:
        if m == ref_method:
            continue
        tt = paired_ttest(results_df, ref_method, m, metric=metric)
        if tt is not None:
            ttests.append(tt)
        d = cohens_d(results_df, ref_method, m, metric=metric)
        if d is not None:
            ds[m] = d

    anova = anova_test(results_df, methods, metric=metric)
    tukey = tukey_hsd_test(results_df, methods, metric=metric)

    return {
        "paired_ttests": ttests,
        "cohens_ds": ds,
        "anova": anova,
        "tukey": tukey,
    }


# =========================================================
# Excel 结果生成
# =========================================================

def create_inner_rrp_excel(
    results_df,
    stat_results,
    optimality_gap_df,
    summary_df,
    problem_config,
    data_config,
    output_path,
):
    """
    创建 Inner RRP 实验的 Excel 报告。
    """
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, LineChart, ScatterChart
    from openpyxl.chart.label import DataLabelList
    from openpyxl.chart.series import DataPoint
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # --- Style definitions ---
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    def style_header_row(ws, row=1, max_col=None):
        if max_col is None:
            max_col = ws.max_column
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = thin_border

    def style_data_table(ws, start_row=1):
        for row in range(start_row, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border
                if row > start_row:
                    cell.alignment = Alignment(horizontal="center")

    def auto_width_columns(ws, min_width=10, max_width=30):
        for col in range(1, ws.max_column + 1):
            max_len = 0
            col_letter = get_column_letter(col)
            for row in range(1, ws.max_row + 1):
                val = str(ws.cell(row=row, column=col).value or "")
                max_len = max(max_len, len(val))
            width = max(min_width, min(max_len + 2, max_width))
            ws.column_dimensions[col_letter].width = width

    # --- Sheet 1: README ---
    ws = wb.active
    ws.title = "README"
    ws.cell(row=1, column=1, value="Inner RRP 基准实验报告")
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value=f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

    info = [
        ("问题参数", ""),
        ("K (每 pack 电芯数)", problem_config.K),
        ("k_max (最大 pack 数)", problem_config.k_max),
        ("delta_bar (方差约束)", problem_config.delta_bar),
        ("w (权重向量)", problem_config.w),
        ("lambda_penalty (方差惩罚)", problem_config.lambda_penalty),
        ("theta1 / theta2 / theta3", f"{problem_config.theta1} / {problem_config.theta2} / {problem_config.theta3}"),
        ("P1 / P2 / P3", f"{problem_config.P1} / {problem_config.P2} / {problem_config.P3}"),
        ("", ""),
        ("数据分布", ""),
        ("mu_C / sigma_C", f"{data_config.mu_C} / {data_config.sigma_C}"),
        ("mu_R / sigma_R", f"{data_config.mu_R} / {data_config.sigma_R}"),
    ]
    for i, (k, v) in enumerate(info, 4):
        ws.cell(row=i, column=1, value=k)
        ws.cell(row=i, column=1).font = Font(bold=True)
        ws.cell(row=i, column=2, value=str(v) if not isinstance(v, (int, float, str, type(None))) else v)

    # --- Sheet 2: Raw_Results ---
    ws2 = wb.create_sheet("Raw_Results")
    raw_cols = ["instance_id", "n_cells", "seed", "method", "reward", "n_packs",
                "avg_delta", "avg_phi", "runtime", "reward_per_pack",
                "positive_pack_ratio", "leftover_count", "P1", "P2", "P3", "P0",
                "gap_to_optimal"]
    for j, col in enumerate(raw_cols, 1):
        ws2.cell(row=1, column=j, value=col)
    for i, row_data in enumerate(results_df[raw_cols].itertuples(index=False), 2):
        for j, val in enumerate(row_data, 1):
            ws2.cell(row=i, column=j, value=val)
    style_header_row(ws2)
    style_data_table(ws2)
    auto_width_columns(ws2)

    # --- Sheet 3: Summary ---
    ws3 = wb.create_sheet("Summary")
    summary_cols = ["method", "n_cells", "mean_reward", "std_reward", "median_reward",
                    "min_reward", "max_reward", "mean_runtime", "mean_n_packs",
                    "mean_reward_per_pack", "mean_positive_ratio", "mean_leftover",
                    "total_P1", "total_P2", "total_P3", "total_P0"]
    for j, col in enumerate(summary_cols, 1):
        ws3.cell(row=1, column=j, value=col)
    for i, row_data in enumerate(summary_df[summary_cols].itertuples(index=False), 2):
        for j, val in enumerate(row_data, 1):
            ws3.cell(row=i, column=j, value=val)
    style_header_row(ws3)
    style_data_table(ws3)
    auto_width_columns(ws3)

    # --- Sheet 4: Statistical_Tests ---
    ws4 = wb.create_sheet("Statistical_Tests")
    row = 1
    ws4.cell(row=row, column=1, value="配对 t 检验 (参考算法: RRP_KMEANS_VNS vs 其他)")
    ws4.cell(row=row, column=1).font = Font(bold=True, size=12)

    tt_headers = ["method_a", "method_b", "mean_a", "mean_b", "mean_diff",
                  "t_statistic", "p_value", "n_pairs", "significant_005", "significant_001"]
    row += 1
    for j, h in enumerate(tt_headers, 1):
        ws4.cell(row=row, column=j, value=h)
    style_header_row(ws4, row=row)

    for tt in stat_results["paired_ttests"]:
        row += 1
        for j, h in enumerate(tt_headers, 1):
            ws4.cell(row=row, column=j, value=tt.get(h, ""))

    row += 2
    ws4.cell(row=row, column=1, value="Cohen's d 效应量 (RRP_KMEANS_VNS vs 其他)")
    ws4.cell(row=row, column=1).font = Font(bold=True, size=12)
    row += 1
    ws4.cell(row=row, column=1, value="method")
    ws4.cell(row=row, column=2, value="cohens_d")
    style_header_row(ws4, row=row, max_col=2)
    for m, d in stat_results["cohens_ds"].items():
        row += 1
        ws4.cell(row=row, column=1, value=m)
        ws4.cell(row=row, column=2, value=round(d, 4))

    if stat_results["anova"]:
        row += 2
        ws4.cell(row=row, column=1, value="ANOVA (整体差异 F 检验)")
        ws4.cell(row=row, column=1).font = Font(bold=True, size=12)
        row += 1
        anova = stat_results["anova"]
        ws4.cell(row=row, column=1, value="F 统计量")
        ws4.cell(row=row, column=2, value=round(anova["f_statistic"], 4))
        ws4.cell(row=row, column=3, value="p 值")
        ws4.cell(row=row, column=4, value=f"{anova['p_value']:.2e}")
        ws4.cell(row=row, column=5, value="显著 (α=0.05)")
        ws4.cell(row=row, column=6, value="是" if anova["significant_005"] else "否")

    if stat_results["tukey"]:
        row += 2
        ws4.cell(row=row, column=1, value="Tukey HSD 事后检验 (成对比较)")
        ws4.cell(row=row, column=1).font = Font(bold=True, size=12)
        row += 1
        tukey_headers = ["method_a", "method_b", "mean_diff", "p_adj", "reject", "conf_low", "conf_high"]
        for j, h in enumerate(tukey_headers, 1):
            ws4.cell(row=row, column=j, value=h)
        style_header_row(ws4, row=row, max_col=len(tukey_headers))
        for tr in stat_results["tukey"]:
            row += 1
            for j, h in enumerate(tukey_headers, 1):
                ws4.cell(row=row, column=j, value=tr.get(h, ""))

    style_data_table(ws4)
    auto_width_columns(ws4)

    # --- Sheet 5: Optimality_Gap ---
    ws5 = wb.create_sheet("Optimality_Gap")
    if optimality_gap_df is not None and len(optimality_gap_df) > 0:
        gap_cols = ["n_cells", "seed", "method", "reward", "optimal_reward", "gap_pct"]
        for j, col in enumerate(gap_cols, 1):
            ws5.cell(row=1, column=j, value=col)
        for i, row_data in enumerate(optimality_gap_df[gap_cols].itertuples(index=False), 2):
            for j, val in enumerate(row_data, 1):
                ws5.cell(row=i, column=j, value=val)
    else:
        ws5.cell(row=1, column=1, value="无小规模实例最优解对比数据")
    style_header_row(ws5)
    style_data_table(ws5)
    auto_width_columns(ws5)

    # --- Sheet 6: ChartData_BoxPlot ---
    ws6 = wb.create_sheet("ChartData_BoxPlot")
    ws6.cell(row=1, column=1, value="method")
    ws6.cell(row=1, column=2, value="n_cells")
    ws6.cell(row=1, column=3, value="reward")
    for i, row_data in enumerate(results_df[["method", "n_cells", "reward"]].itertuples(index=False), 2):
        ws6.cell(row=i, column=1, value=row_data[0])
        ws6.cell(row=i, column=2, value=row_data[1])
        ws6.cell(row=i, column=3, value=row_data[2])
    style_header_row(ws6)
    auto_width_columns(ws6)

    # --- Sheet 7: ChartData_Runtime ---
    ws7 = wb.create_sheet("ChartData_Runtime")
    ws7.cell(row=1, column=1, value="method")
    ws7.cell(row=1, column=2, value="n_cells")
    ws7.cell(row=1, column=3, value="runtime")
    for i, row_data in enumerate(results_df[["method", "n_cells", "runtime"]].itertuples(index=False), 2):
        ws7.cell(row=i, column=1, value=row_data[0])
        ws7.cell(row=i, column=2, value=row_data[1])
        ws7.cell(row=i, column=3, value=row_data[2])
    style_header_row(ws7)
    auto_width_columns(ws7)

    # --- Sheet 8: Charts (使用 matplotlib 生成图表图片) ---
    ws8 = wb.create_sheet("Charts")
    ws8.cell(row=1, column=1, value="图表将在运行后由 matplotlib 生成，请参考单独的图表文件。")
    ws8.cell(row=2, column=1, value="或者使用 Excel 的插入图表功能，基于 ChartData_* sheets 的数据手动创建图表。")

    # Save
    os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else ".", exist_ok=True)
    wb.save(output_path)
    print(f"Excel 报告已保存至: {output_path}")


def create_two_stage_excel(
    results_df,
    summary_df,
    problem_config,
    data_config,
    output_path,
):
    """
    创建 Two-Stage 实验的 Excel 报告。
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    def style_header_row(ws, row=1, max_col=None):
        if max_col is None:
            max_col = ws.max_column
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = thin_border

    def style_data_table(ws, start_row=1):
        for row in range(start_row, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border
                if row > start_row:
                    cell.alignment = Alignment(horizontal="center")

    def auto_width_columns(ws, min_width=10, max_width=30):
        for col in range(1, ws.max_column + 1):
            max_len = 0
            col_letter = get_column_letter(col)
            for row in range(1, ws.max_row + 1):
                val = str(ws.cell(row=row, column=col).value or "")
                max_len = max(max_len, len(val))
            width = max(min_width, min(max_len + 2, max_width))
            ws.column_dimensions[col_letter].width = width

    # --- README ---
    ws = wb.active
    ws.title = "README"
    ws.cell(row=1, column=1, value="Two-Stage 完整系统实验报告")
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value=f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

    # --- Raw_Results ---
    ws2 = wb.create_sheet("Raw_Results")
    cols = list(results_df.columns)
    for j, col in enumerate(cols, 1):
        ws2.cell(row=1, column=j, value=col)
    for i, row_data in enumerate(results_df.itertuples(index=False), 2):
        for j, val in enumerate(row_data, 1):
            ws2.cell(row=i, column=j, value=val)
    style_header_row(ws2)
    style_data_table(ws2)
    auto_width_columns(ws2)

    # --- Summary ---
    ws3 = wb.create_sheet("Summary")
    cols = list(summary_df.columns)
    for j, col in enumerate(cols, 1):
        ws3.cell(row=1, column=j, value=col)
    for i, row_data in enumerate(summary_df.itertuples(index=False), 2):
        for j, val in enumerate(row_data, 1):
            ws3.cell(row=i, column=j, value=val)
    style_header_row(ws3)
    style_data_table(ws3)
    auto_width_columns(ws3)

    os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else ".", exist_ok=True)
    wb.save(output_path)
    print(f"Excel 报告已保存至: {output_path}")
